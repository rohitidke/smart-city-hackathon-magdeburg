#!/usr/bin/env python3
"""
Parse Stabstelle Klima Excel files to JSON
"""
import json
import openpyxl
from pathlib import Path


def parse_energy_emissions():
    """Parse Energieverbrauch und Emissionen CO2_1.xlsx"""
    file = Path(__file__).parent / "Energieverbrauch und Emissionen CO2_1.xlsx"
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb.active

    # Row 2 has years starting from column C (index 2)
    years_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    years = [int(y) for y in years_row[2:] if y is not None]

    # Parse each sector
    sectors = {}
    sector_rows = {
        "Haushalte": 3,
        "Industrie": 4,
        "GHD": 5,  # Gewerbe, Handel, Dienstleistungen
        "Kommunale Verwaltung": 6,
        "Verkehr": 7,
        "Kommunale Flotte": 8,
    }

    result = []
    for year_idx, year in enumerate(years):
        row_data = {"year": year}
        for sector, row_num in sector_rows.items():
            values = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            value = values[2 + year_idx]  # Skip first 2 columns
            if value is not None:
                row_data[sector.lower().replace(" ", "_")] = round(float(value), 2)

        # Calculate total
        row_data["total"] = round(sum(v for k, v in row_data.items() if k != "year"), 2)
        result.append(row_data)

    return result


def parse_led_conversion():
    """Parse LED-Umrüstung Straßenlaternen.xlsx"""
    file = Path(__file__).parent / "LED-Umrüstung Straßenlaternen.xlsx"
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb.active

    # Extract values from specific cells
    total_lights = int(sheet.cell(1, 8).value) if sheet.cell(1, 8).value else 0
    led_converted = int(sheet.cell(2, 8).value) if sheet.cell(2, 8).value else 0

    return {
        "total_streetlights": total_lights,
        "led_converted": led_converted,
        "conventional": total_lights - led_converted,
        "led_percentage": round((led_converted / total_lights * 100), 1) if total_lights > 0 else 0,
    }


def parse_solar_energy():
    """Parse Strom aus solare Strahlungsenergie_1.xlsx"""
    file = Path(__file__).parent / "Strom aus solare Strahlungsenergie_1.xlsx"
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb.active

    # Extract PV installation data from rows 3-6
    # Row 6 has totals
    total_row = list(sheet.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    stats_row = list(sheet.iter_rows(min_row=8, max_row=8, values_only=True))[0]

    return {
        "total_installations": int(total_row[1]) if total_row[1] else 0,
        "installed_capacity_kw": round(float(total_row[2]), 2) if total_row[2] else 0,
        "installed_capacity_mw": round(float(total_row[2]) / 1000, 2) if total_row[2] else 0,
        "annual_generation_mwh": int(stats_row[4]) if stats_row[4] else 0,
        "annual_generation_gwh": round(float(stats_row[5]), 2) if stats_row[5] else 0,
    }


def main():
    output_dir = Path(__file__).parent

    # Parse all datasets
    print("Parsing Stabstelle Klima data...")

    energy = parse_energy_emissions()
    with open(output_dir / "energy_emissions.json", "w", encoding="utf-8") as f:
        json.dump(energy, f, indent=2, ensure_ascii=False)
    print(f"✓ Wrote energy_emissions.json ({len(energy)} years)")

    led = parse_led_conversion()
    with open(output_dir / "led_streetlights.json", "w", encoding="utf-8") as f:
        json.dump(led, f, indent=2, ensure_ascii=False)
    print(f"✓ Wrote led_streetlights.json")

    solar = parse_solar_energy()
    with open(output_dir / "solar_energy.json", "w", encoding="utf-8") as f:
        json.dump(solar, f, indent=2, ensure_ascii=False)
    print(f"✓ Wrote solar_energy.json")

    print("\nDone! Climate data parsed successfully.")


if __name__ == "__main__":
    main()
